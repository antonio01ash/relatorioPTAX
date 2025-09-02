import requests
import datetime
import pandas as pd
from openpyxl import load_workbook

def ultimo_dia_util():
    hoje = datetime.date.today()
    dia = hoje - datetime.timedelta(days=1)
    while dia.weekday() >= 5:
        dia -= datetime.timedelta(days=1)
    return dia

def pegar_ptax(data):
    data_formatada = data.strftime("%m-%d-%Y")
    url = f"https://olinda.bcb.gov.br/olinda/servico/PTAX/versao/v1/odata/CotacaoDolarDia(dataCotacao=@dataCotacao)?@dataCotacao='{data_formatada}'&$top=100&$format=json"
    res = requests.get(url).json()
    df = pd.DataFrame(res["value"])
    if df.empty:
        return None
    ultima_cotacao = df.sort_values("horaCotacao").iloc[-1]
    return {
        "Data": data,
        "Compra (PTAX)": ultima_cotacao["cotacaoCompra"],
        "Venda (PTAX)": ultima_cotacao["cotacaoVenda"]
    }

dia_util = ultimo_dia_util()
cotacao = pegar_ptax(dia_util)

if cotacao:
    relatorio = pd.DataFrame([cotacao])
    arquivo = "relatorio_ptax.xlsx"

    try:
        book = load_workbook(arquivo)
        writer = pd.ExcelWriter(arquivo, engine="openpyxl")
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}
        startrow = writer.sheets["Sheet1"].max_row
        relatorio.to_excel(writer, index=False, header=False, startrow=startrow)
        writer.close()
    except FileNotFoundError:
        relatorio.to_excel(arquivo, index=False)

    print("✅ Relatório atualizado")
else:
    print("❌ Não foi encontrada cotação para o dia útil anterior.")
